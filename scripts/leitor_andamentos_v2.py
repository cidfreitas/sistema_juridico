#!/usr/bin/env python3
"""
leitor_andamentos_v2.py - Versão corrigida
Extrai corretamente números CNJ dos hyperlinks
"""

import openpyxl
import json
import re
from datetime import datetime
from pathlib import Path

CONFIG_FILE = "/Users/cidfreitas/Documents/scraper_tjsp_diario/data_scraping.xlsx"
ANDAMENTOS_FILE = "/Users/cidfreitas/Documents/scraper_tjsp_diario/andamentos_diario_tjsp.xlsx"
OUTPUT_FILE = "../data/saida/processos.json"

print("=" * 70)
print("LEITOR DE ANDAMENTOS V2 - CORRIGIDO")
print("=" * 70)

# 1. LER CONFIGURAÇÃO
print("\n📖 Lendo data_scraping.xlsx...")
wb_config = openpyxl.load_workbook(CONFIG_FILE)
ws_config = wb_config.active

processos_config = {}
for row in ws_config.iter_rows(min_row=2):
    if row[3] and row[3].value:  # Se tem número de processo
        cnj = str(row[3].value).strip()
        pasta = row[0].value if row[0].value else "N/A"
        parte = row[4].value if row[4].value else "N/A"
        
        processos_config[cnj] = {
            "pasta": pasta,
            "parte": parte
        }

print(f"✅ {len(processos_config)} processos configurados")

# 2. LER ANDAMENTOS
print("\n📖 Lendo andamentos_diario_tjsp.xlsx...")
wb_andamentos = openpyxl.load_workbook(ANDAMENTOS_FILE)
ws_andamentos = wb_andamentos.active

def extrair_cnj(valor):
    """Extrai número CNJ de hyperlink ou texto"""
    if not valor:
        return None
    
    valor_str = str(valor).strip()
    
    # Se é hyperlink do Excel: =HYPERLINK(...,"NÚMERO")
    match = re.search(r'"(\d{7}-\d{2}\.\d{4}\.\d{1}\.\d{2}\.\d{4})"', valor_str)
    if match:
        return match.group(1)
    
    # Se é texto direto
    match = re.search(r'(\d{7}-\d{2}\.\d{4}\.\d{1}\.\d{2}\.\d{4})', valor_str)
    if match:
        return match.group(1)
    
    return None

andamentos_por_processo = {}
for row in ws_andamentos.iter_rows(min_row=2):
    ref = row[0].value
    numero_raw = row[1].value
    partes = row[2].value
    origem = row[3].value
    data_mov = row[4].value
    descricao = row[5].value
    
    # Extrair CNJ
    cnj = extrair_cnj(numero_raw)
    
    if not cnj:
        continue
    
    if cnj not in andamentos_por_processo:
        andamentos_por_processo[cnj] = {
            "ref": ref,
            "partes": partes or "N/A",
            "origem": origem or "N/A",
            "andamentos": []
        }
    
    andamentos_por_processo[cnj]["andamentos"].append({
        "data": data_mov,
        "descricao": descricao
    })

print(f"✅ {len(andamentos_por_processo)} processos com andamentos")

# 3. MESCLAR E ESTRUTURAR
print("\n🔧 Estruturando dados...")
processos_finais = []

for cnj, andamentos_info in andamentos_por_processo.items():
    config_info = processos_config.get(cnj, {})
    
    # Determinar status
    status = "novo"
    if andamentos_info["andamentos"]:
        ultima_data = andamentos_info["andamentos"][-1]["data"]
        if ultima_data:
            try:
                dias_desde = (datetime.now() - datetime.strptime(str(ultima_data), "%d-%m-%Y")).days
                if dias_desde <= 3:
                    status = "urgente"
                elif dias_desde <= 10:
                    status = "andamento"
                elif dias_desde <= 30:
                    status = "espera"
            except:
                status = "andamento"
    
    processo = {
        "cnj": cnj,
        "status": status,
        "pasta": config_info.get("pasta", "N/A"),
        "parte": config_info.get("parte", "N/A"),
        "partes_processo": andamentos_info["partes"],
        "origem": andamentos_info["origem"],
        "total_andamentos": len(andamentos_info["andamentos"]),
        "ultimo_andamento": andamentos_info["andamentos"][-1] if andamentos_info["andamentos"] else None,
        "andamentos": andamentos_info["andamentos"][-5:] if len(andamentos_info["andamentos"]) > 5 else andamentos_info["andamentos"],
    }
    
    processos_finais.append(processo)

print(f"✅ {len(processos_finais)} processos estruturados")

# 4. SALVAR JSON
print(f"\n💾 Salvando em {OUTPUT_FILE}...")
Path(OUTPUT_FILE).parent.mkdir(parents=True, exist_ok=True)

dados = {
    "metadata": {
        "data_exportacao": datetime.now().isoformat(),
        "total_processos": len(processos_finais),
        "total_andamentos": sum(p["total_andamentos"] for p in processos_finais),
        "version": "2.0"
    },
    "processos": processos_finais
}

with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    json.dump(dados, f, indent=2, ensure_ascii=False)

print(f"✅ JSON criado!\n")

# 5. RESUMO
print("=" * 70)
print("RESUMO")
print("=" * 70)
urgentes = sum(1 for p in processos_finais if p["status"] == "urgente")
andamento = sum(1 for p in processos_finais if p["status"] == "andamento")
espera = sum(1 for p in processos_finais if p["status"] == "espera")
novo = sum(1 for p in processos_finais if p["status"] == "novo")

print(f"🔴 URGENTES: {urgentes}")
print(f"🟡 EM ANDAMENTO: {andamento}")
print(f"⏸️  EM ESPERA: {espera}")
print(f"🟦 NOVOS: {novo}")
print(f"\n📊 TOTAL: {len(processos_finais)} processos")
print(f"📈 ANDAMENTOS: {sum(p['total_andamentos'] for p in processos_finais)}")
print(f"\n✅ Arquivo: {OUTPUT_FILE}")

